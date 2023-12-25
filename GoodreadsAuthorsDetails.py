import openpyxl
import requests
from bs4 import BeautifulSoup
from AuthorScrapper import AuthorScrapper
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet['A1'] = 'Author name'
worksheet['B1'] = 'Website'
worksheet['C1'] = 'Twitter'
worksheet['D1'] = 'Rating'
worksheet['E1'] = 'Book details'
row = 2
url = "https://www.goodreads.com/author/on_goodreads?btnSubmit=Search&page="
pageno = 1
while True:
    try:
        urls = url + str(pageno)
        response = requests.get(urls)
        soup = BeautifulSoup(response.content, "html.parser")
        authorids = soup.find_all("a", class_="bookAuthorProfile__name", href=True)
        for authorid in authorids:
            authorURL = "https://www.goodreads.com/"+authorid["href"]
            scrapper = AuthorScrapper.scrap(authorURL)
            worksheet.cell(row=row, column=1, value=scrapper[0])
            worksheet.cell(row=row, column=2, value=scrapper[1])
            worksheet.cell(row=row, column=3, value=scrapper[2])
            worksheet.cell(row=row, column=4, value=float(scrapper[3]) if scrapper[3]!="NA" else "NA")
            worksheet.cell(row=row, column=5, value=scrapper[4])
            print(row)
            row+=1
        next = soup.find("a", {"class": "next_page"})
        if not next:
            break
        pageno+=1
    except Exception as ex:
        print(ex)
workbook.save('goodreads_authors.xlsx')